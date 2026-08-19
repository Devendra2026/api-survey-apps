#!/usr/bin/env python3
"""Generate survey_data_modern.xlsx — municipal survey data-entry template."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Protection, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.table import TableStyle, TableStyleElement
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "survey_data_modern.xlsx"

# Design system (Data-Dense Dashboard)
PRIMARY = "1E40AF"
SECONDARY = "3B82F6"
ACCENT = "F59E0B"
BACKGROUND = "F8FAFC"
WHITE = "FFFFFF"
BORDER_COLOR = "D3D3D3"
TEXT = "1E3A8A"
MUTED = "64748B"

STATUS_FILLS = {
    "Completed": "E8F8F0",
    "Draft": "FFF8E1",
    "Submitted": "EAF4FC",
    "Approved": "F4EAF8",
}
STATUS_FONTS = {
    "Completed": "1E8449",
    "Draft": "B7950B",
    "Submitted": "1A5276",
    "Approved": "6C3483",
}

HEADERS = [
    "SN",
    "Status",
    "Surveyor Name",
    "Assessment Year",
    "Property Id",
    "Date of Survey",
    "Owner Name",
    "Owner Father Name",
    "Mobile No",
    "Ward Name",
    "Is Slum",
    "Parcel No",
    "Unit No",
    "old property Number",
    "Constructed Date",
    "Respondent Name",
    "Respondent Relationship",
    "City",
    "Pincode",
    "House No",
    "Locality",
    "Colony",
    "Tax Rate Zone",
    "Property Ownership",
    "Property Type",
    "Property Uses",
    "Situation",
    "Road Type",
    "Floors",
    "Plot Area SqFt",
    "Plinth Area SqFt",
    "Total Built Up Area SqFt",
    "Is Municipal Water Supply",
    "Total Water Connection",
    "Water Connection Id/Type",
    "Toilet Type",
    "Is Municipal Waste Service",
    "Latitude",
    "Longitude",
]

assert len(HEADERS) == 39

# 1-based groups for visual separators
GROUPS = [
    (1, 6, "Survey / Workflow", PRIMARY),
    (7, 17, "Owner / Respondent", "1E3A5F"),
    (18, 22, "Address Information", "1D4ED8"),
    (23, 29, "Property Classification", "1E40AF"),
    (30, 32, "Area Measurements", "047857"),
    (33, 37, "Utility Information", "1E3A8A"),
    (38, 39, "GPS / Location", ACCENT),
]
GROUP_STARTS = {start for start, _, _, _ in GROUPS}

NAMED_LISTS: dict[str, list[str]] = {
    "StatusOptions": ["Completed", "Draft", "Submitted", "Approved"],
    "AssessmentYearOptions": ["2025-2026", "2026-2027", "2024-2025"],
    "SlumOptions": ["yes", "no"],
    "RelationshipOptions": [
        "Self",
        "Son",
        "Daughter",
        "Wife",
        "Husband",
        "Mother",
        "Father",
        "Sister",
        "Brother",
        "Neighbour",
        "Other",
        "N/A",
    ],
    "TaxRateZoneOptions": ["RATE ZONE 1", "RATE ZONE 2", "RATE ZONE 3", "RATE ZONE 4"],
    "OwnershipOptions": [
        "Individual (Single/Joint)",
        "Limited Company",
        "Trust Society",
        "Religious Body",
        "Joint",
    ],
    "PropertyTypeOptions": [
        "Residential",
        "Commercial",
        "Open Land",
        "Mix",
        "Mandir",
        "School College",
        "Hospital Nursing Pathology",
        "Mall Showroom",
        "Shop Bakery",
        "Godown",
        "Industrial/Factory",
        "Hotel/Restaurant/Marriage Garden",
        "Bakery",
    ],
    "PropertyUseOptions": [
        "Residential Self",
        "Residential Rented",
        "Open Land",
        "Agriculture",
        "Shops/Banks",
        "Industrial/Factory",
        "Bakery",
        "Hotel/Restaurant/Marriage Garden",
        "Religious Property",
        "Residential/Commercial",
        "Commercial",
        "Mix Property",
        "Mandir",
        "Godown",
        "School College",
        "Hospital Nursing Pathology",
    ],
    "SituationOptions": ["Main Road", "Main Market", "Interior", "Slum"],
    "RoadTypeOptions": [
        "Dambar Road",
        "Kachcha Road",
        "RCC Road",
        "Interlocking Road",
        "Rcc",
        "Kaccha",
    ],
    "WaterSupplyOptions": ["yes", "no"],
    "WaterConnectionTypeOptions": [
        "null",
        "Government Tap",
        "Borewell",
        "Dug well",
        "Other",
        "Sewer System",
        "Septic Tank",
        "Surface Drain",
        "No Toilet",
        "Dry/Bucket Latrine",
        "Connected to Specific Tank",
        "Connected to Surface Drains",
        "Pour-flush pit latrine",
    ],
    "ToiletTypeOptions": [
        "No Toilet",
        "Connected to Specific Tank",
        "Connected to Surface Drains",
        "Dry/Bucket Latrine",
        "Pour-flush pit latrine",
        "Other",
        "Septic Tank",
        "Sewer System",
    ],
    "WasteServiceOptions": ["yes", "no"],
    "ActionOptions": ["Demand Notice", "Print", "View", "Edit", "Delete"],
}

TITLE_ROW = 1
META_ROW = 2
GROUP_ROW = 3
HEADER_ROW = 4
FIRST_DATA_ROW = 5
SAMPLE_COUNT = 15
EXTRA_ROWS = 65
TOTAL_DATA_ROWS = SAMPLE_COUNT + EXTRA_ROWS
LAST_DATA_ROW = HEADER_ROW + TOTAL_DATA_ROWS  # 84
TABLE_NAME = "SurveyData"
EDITABLE_FIELD_COUNT = 37  # all columns except SN and Property Id

THIN = Side(style="thin", color=BORDER_COLOR)
MED_ACCENT = Side(style="medium", color=ACCENT)
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_NAME = "Calibri"


def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def font(**kwargs: object) -> Font:
    kwargs.setdefault("name", FONT_NAME)
    kwargs.setdefault("color", TEXT)
    return Font(**kwargs)  # type: ignore[arg-type]


def comment(text: str, width: int = 280, height: int = 90) -> Comment:
    note = Comment(text, "Survey Template")
    note.width = width
    note.height = height
    return note


def col_letter(index: int) -> str:
    return get_column_letter(index)


def data_range(col_idx: int) -> str:
    return f"{col_letter(col_idx)}{FIRST_DATA_ROW}:{col_letter(col_idx)}{LAST_DATA_ROW}"


def table_ref() -> str:
    return f"A{HEADER_ROW}:{col_letter(len(HEADERS))}{LAST_DATA_ROW}"


def apply_border(cell, left_accent: bool = False) -> None:
    cell.border = Border(
        left=MED_ACCENT if left_accent else THIN,
        right=THIN,
        top=THIN,
        bottom=THIN,
    )


def register_named_style(wb: Workbook) -> None:
    header_style = NamedStyle(name="SurveyHeader")
    header_style.font = font(bold=True, color=WHITE, size=11)
    header_style.fill = fill(PRIMARY)
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = THIN_BORDER
    if "SurveyHeader" not in wb.named_styles:
        wb.add_named_style(header_style)


def register_custom_table_style(wb: Workbook) -> str:
    style_name = "SurveyModern"
    header_dxf = DifferentialStyle(
        font=Font(bold=True, color=WHITE, name=FONT_NAME, size=11),
        fill=PatternFill(patternType="solid", fgColor=PRIMARY),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=THIN_BORDER,
    )
    stripe1 = DifferentialStyle(fill=PatternFill(patternType="solid", fgColor=WHITE))
    stripe2 = DifferentialStyle(fill=PatternFill(patternType="solid", fgColor=BACKGROUND))
    dss = wb._differential_styles
    header_id = len(dss.styles)
    dss.styles.append(header_dxf)
    s1_id = len(dss.styles)
    dss.styles.append(stripe1)
    s2_id = len(dss.styles)
    dss.styles.append(stripe2)
    ts = TableStyle(name=style_name, pivot=False)
    ts.tableStyleElement.append(TableStyleElement(type="headerRow", size=1, dxfId=header_id))
    ts.tableStyleElement.append(TableStyleElement(type="firstRowStripe", size=1, dxfId=s1_id))
    ts.tableStyleElement.append(TableStyleElement(type="secondRowStripe", size=1, dxfId=s2_id))
    styles = list(wb._table_styles.tableStyle)
    styles.append(ts)
    wb._table_styles.tableStyle = styles
    return style_name


def sample_rows() -> list[dict[str, object]]:
    """Fictional municipal survey rows. SN and Property Id are formula-driven."""
    return [
        {
            "Status": "Completed",
            "Surveyor Name": "Amit Verma",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 6, 12, 10, 35),
            "Owner Name": "Ramesh Kumar Sharma",
            "Owner Father Name": "Ram Lal Sharma",
            "Mobile No": "9876501234",
            "Ward Name": "Ward 11 - Gandhi Nagar",
            "Is Slum": "no",
            "Parcel No": 125,
            "Unit No": 1,
            "old property Number": 4587,
            "Constructed Date": date(2008, 3, 14),
            "Respondent Name": "Ramesh Kumar Sharma",
            "Respondent Relationship": "Self",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "12-A",
            "Locality": "Gandhi Nagar",
            "Colony": "Shastri Colony",
            "Tax Rate Zone": "RATE ZONE 1",
            "Property Ownership": "Individual (Single/Joint)",
            "Property Type": "Residential",
            "Property Uses": "Residential Self",
            "Situation": "Main Road",
            "Road Type": "Dambar Road",
            "Floors": 2,
            "Plot Area SqFt": 1200.5,
            "Plinth Area SqFt": 980.25,
            "Total Built Up Area SqFt": 1760.75,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Sewer System",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5583,
            "Longitude": 78.6628,
        },
        {
            "Status": "Draft",
            "Surveyor Name": "Neha Singh",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 7, 3, 16, 10),
            "Owner Name": "Sunrise Traders Pvt Ltd",
            "Owner Father Name": "N/A",
            "Mobile No": "9123456780",
            "Ward Name": "Ward 3 - Nehru Market",
            "Is Slum": "no",
            "Parcel No": 210,
            "Unit No": 2,
            "old property Number": 2201,
            "Constructed Date": date(2015, 11, 2),
            "Respondent Name": "Sanjay Kapoor",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "Shop 4",
            "Locality": "Nehru Market",
            "Colony": "Market Road",
            "Tax Rate Zone": "RATE ZONE 2",
            "Property Ownership": "Limited Company",
            "Property Type": "Commercial",
            "Property Uses": "Shops/Banks",
            "Situation": "Main Market",
            "Road Type": "RCC Road",
            "Floors": 1,
            "Plot Area SqFt": 850.0,
            "Plinth Area SqFt": 820.5,
            "Total Built Up Area SqFt": 820.5,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 2,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Septic Tank",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5575,

            "Longitude": 78.6631,
        },
        {
            "Status": "Submitted",
            "Surveyor Name": "Pooja Yadav",
            "Assessment Year": "2026-2027",
            "Date of Survey": datetime(2026, 4, 18, 9, 20),
            "Owner Name": "Anita Gupta",
            "Owner Father Name": "Mahesh Gupta",
            "Mobile No": "9988776655",
            "Ward Name": "Ward 7 - Subhash Nagar",
            "Is Slum": "no",
            "Parcel No": 88,
            "Unit No": 1,
            "old property Number": 3104,
            "Constructed Date": date(2010, 8, 21),
            "Respondent Name": "Rohit Gupta",
            "Respondent Relationship": "Son",
            "City": "Etah",
            "Pincode": "207002",
            "House No": "44/2",
            "Locality": "Subhash Nagar",
            "Colony": "Gupta Colony",
            "Tax Rate Zone": "RATE ZONE 3",
            "Property Ownership": "Joint",
            "Property Type": "Mix",
            "Property Uses": "Mix Property",
            "Situation": "Interior",
            "Road Type": "Interlocking Road",
            "Floors": 3,
            "Plot Area SqFt": 1500.0,
            "Plinth Area SqFt": 1100.0,
            "Total Built Up Area SqFt": 2650.4,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 2,
            "Water Connection Id/Type": "Borewell",
            "Toilet Type": "Sewer System",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5591,

            "Longitude": 78.6642,
        },
        {
            "Status": "Approved",
            "Surveyor Name": "Imran Khan",
            "Assessment Year": "2024-2025",
            "Date of Survey": datetime(2025, 2, 9, 11, 45),
            "Owner Name": "Vikram Singh",
            "Owner Father Name": "Gopal Singh",
            "Mobile No": "9012345678",
            "Ward Name": "Ward 1 - Civil Lines",
            "Is Slum": "no",
            "Parcel No": 45,
            "Unit No": 1,
            "old property Number": 1180,
            "Constructed Date": date(1998, 1, 5),
            "Respondent Name": "Vikram Singh",
            "Respondent Relationship": "Self",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "Plot 9",
            "Locality": "Civil Lines",
            "Colony": "Officer Colony",
            "Tax Rate Zone": "RATE ZONE 4",
            "Property Ownership": "Individual (Single/Joint)",
            "Property Type": "Open Land",
            "Property Uses": "Open Land",
            "Situation": "Interior",
            "Road Type": "Kachcha Road",
            "Floors": 0,
            "Plot Area SqFt": 2400.0,
            "Plinth Area SqFt": 0.0,
            "Total Built Up Area SqFt": 0.0,
            "Is Municipal Water Supply": "no",
            "Total Water Connection": 0,
            "Water Connection Id/Type": "null",
            "Toilet Type": "No Toilet",
            "Is Municipal Waste Service": "no",
            "Latitude": 27.5578,

            "Longitude": 78.6612,
        },
        {
            "Status": "Completed",
            "Surveyor Name": "Amit Verma",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 8, 22, 14, 5),
            "Owner Name": "Sunita Devi",
            "Owner Father Name": "Hari Prasad",
            "Mobile No": "9765432109",
            "Ward Name": "Ward 11 - Gandhi Nagar",
            "Is Slum": "no",
            "Parcel No": 312,
            "Unit No": 3,
            "old property Number": 5612,
            "Constructed Date": date(2018, 6, 30),
            "Respondent Name": "Sunita Devi",
            "Respondent Relationship": "Wife",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "7/B",
            "Locality": "Bazaar Road",
            "Colony": "New Market",
            "Tax Rate Zone": "RATE ZONE 1",
            "Property Ownership": "Individual (Single/Joint)",
            "Property Type": "Shop Bakery",
            "Property Uses": "Bakery",
            "Situation": "Main Market",
            "Road Type": "Dambar Road",
            "Floors": 1,
            "Plot Area SqFt": 420.75,
            "Plinth Area SqFt": 400.5,
            "Total Built Up Area SqFt": 400.5,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Connected to Specific Tank",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5589,

            "Longitude": 78.6654,
        },
        {
            "Status": "Draft",
            "Surveyor Name": "Kavita Joshi",
            "Assessment Year": "2026-2027",
            "Date of Survey": datetime(2026, 5, 2, 12, 40),
            "Owner Name": "Saraswati Education Trust",
            "Owner Father Name": "N/A",
            "Mobile No": "8899001122",
            "Ward Name": "Ward 5 - Jawahar Nagar",
            "Is Slum": "no",
            "Parcel No": 67,
            "Unit No": 1,
            "old property Number": 880,
            "Constructed Date": date(2001, 9, 12),
            "Respondent Name": "Meena Kumari",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207003",
            "House No": "School Campus",
            "Locality": "Jawahar Nagar",
            "Colony": "Education Hub",
            "Tax Rate Zone": "RATE ZONE 2",
            "Property Ownership": "Trust Society",
            "Property Type": "School College",
            "Property Uses": "School College",
            "Situation": "Main Road",
            "Road Type": "RCC Road",
            "Floors": 3,
            "Plot Area SqFt": 18000.0,
            "Plinth Area SqFt": 12500.0,
            "Total Built Up Area SqFt": 28450.5,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 4,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Sewer System",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5562,

            "Longitude": 78.6608,
        },
        {
            "Status": "Submitted",
            "Surveyor Name": "Ravi Chauhan",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 9, 14, 17, 15),
            "Owner Name": "Abdul Karim",
            "Owner Father Name": "Mohammed Yusuf",
            "Mobile No": "9345678123",
            "Ward Name": "Ward 9 - Basti Area",
            "Is Slum": "yes",
            "Parcel No": 401,
            "Unit No": 1,
            "old property Number": 102,
            "Constructed Date": date(1995, 4, 8),
            "Respondent Name": "Farhan Ali",
            "Respondent Relationship": "Neighbour",
            "City": "Etah",
            "Pincode": "207002",
            "House No": "Hut 18",
            "Locality": "Basti Area",
            "Colony": "Ambedkar Basti",
            "Tax Rate Zone": "RATE ZONE 4",
            "Property Ownership": "Individual (Single/Joint)",
            "Property Type": "Residential",
            "Property Uses": "Residential Rented",
            "Situation": "Slum",
            "Road Type": "Kaccha",
            "Floors": 1,
            "Plot Area SqFt": 360.0,
            "Plinth Area SqFt": 320.0,
            "Total Built Up Area SqFt": 320.0,
            "Is Municipal Water Supply": "no",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Dug well",
            "Toilet Type": "Dry/Bucket Latrine",
            "Is Municipal Waste Service": "no",
            "Latitude": 27.5554,

            "Longitude": 78.6671,
        },
        {
            "Status": "Approved",
            "Surveyor Name": "Neha Singh",
            "Assessment Year": "2024-2025",
            "Date of Survey": datetime(2025, 1, 28, 13, 55),
            "Owner Name": "Royal Banquet Pvt Ltd",
            "Owner Father Name": "N/A",
            "Mobile No": "9811122233",
            "Ward Name": "Ward 2 - Station Road",
            "Is Slum": "no",
            "Parcel No": 156,
            "Unit No": 1,
            "old property Number": 7741,
            "Constructed Date": date(2019, 12, 1),
            "Respondent Name": "Priya Verma",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "NH-34, Plot 6",
            "Locality": "Station Road",
            "Colony": "Highway Complex",
            "Tax Rate Zone": "RATE ZONE 1",
            "Property Ownership": "Limited Company",
            "Property Type": "Hotel/Restaurant/Marriage Garden",
            "Property Uses": "Hotel/Restaurant/Marriage Garden",
            "Situation": "Main Road",
            "Road Type": "Dambar Road",
            "Floors": 2,
            "Plot Area SqFt": 9200.0,
            "Plinth Area SqFt": 4800.0,
            "Total Built Up Area SqFt": 7600.25,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 3,
            "Water Connection Id/Type": "Borewell",
            "Toilet Type": "Sewer System",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5602,

            "Longitude": 78.6589,
        },
        {
            "Status": "Completed",
            "Surveyor Name": "Imran Khan",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 10, 5, 8, 50),
            "Owner Name": "Shri Ram Mandir Committee",
            "Owner Father Name": "N/A",
            "Mobile No": "9456123780",
            "Ward Name": "Ward 8 - Mandir Ward",
            "Is Slum": "no",
            "Parcel No": 22,
            "Unit No": 1,
            "old property Number": 19,
            "Constructed Date": date(1972, 2, 18),
            "Respondent Name": "Suresh Chandra",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "Mandir Complex",
            "Locality": "Old Town",
            "Colony": "Mandir Gali",
            "Tax Rate Zone": "RATE ZONE 3",
            "Property Ownership": "Religious Body",
            "Property Type": "Mandir",
            "Property Uses": "Mandir",
            "Situation": "Interior",
            "Road Type": "Interlocking Road",
            "Floors": 1,
            "Plot Area SqFt": 3100.0,
            "Plinth Area SqFt": 1450.0,
            "Total Built Up Area SqFt": 1450.0,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Other",
            "Toilet Type": "Pour-flush pit latrine",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5586,

            "Longitude": 78.6638,
        },
        {
            "Status": "Draft",
            "Surveyor Name": "Pooja Yadav",
            "Assessment Year": "2026-2027",
            "Date of Survey": datetime(2026, 4, 27, 15, 25),
            "Owner Name": "Agro Mill Industries",
            "Owner Father Name": "N/A",
            "Mobile No": "9001122334",
            "Ward Name": "Ward 4 - Industrial Area",
            "Is Slum": "no",
            "Parcel No": 890,
            "Unit No": 5,
            "old property Number": 6402,
            "Constructed Date": date(2012, 7, 19),
            "Respondent Name": "Deepak Yadav",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207003",
            "House No": "Shed 11",
            "Locality": "Industrial Area",
            "Colony": "SIDCUL Cluster",
            "Tax Rate Zone": "RATE ZONE 2",
            "Property Ownership": "Limited Company",
            "Property Type": "Industrial/Factory",
            "Property Uses": "Industrial/Factory",
            "Situation": "Main Road",
            "Road Type": "RCC Road",
            "Floors": 1,
            "Plot Area SqFt": 14500.0,
            "Plinth Area SqFt": 9800.0,
            "Total Built Up Area SqFt": 11200.8,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 3,
            "Water Connection Id/Type": "Borewell",
            "Toilet Type": "Septic Tank",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5571,

            "Longitude": 78.6663,
        },
        {
            "Status": "Submitted",
            "Surveyor Name": "Kavita Joshi",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 11, 11, 10, 5),
            "Owner Name": "City Care Hospital Trust",
            "Owner Father Name": "N/A",
            "Mobile No": "9877001122",
            "Ward Name": "Ward 6 - Hospital Road",
            "Is Slum": "no",
            "Parcel No": 333,
            "Unit No": 1,
            "old property Number": 2500,
            "Constructed Date": date(2016, 5, 9),
            "Respondent Name": "Dr. Rajesh Patel",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "Plot 21",
            "Locality": "Hospital Road",
            "Colony": "Medical Campus",
            "Tax Rate Zone": "RATE ZONE 1",
            "Property Ownership": "Trust Society",
            "Property Type": "Hospital Nursing Pathology",
            "Property Uses": "Hospital Nursing Pathology",
            "Situation": "Main Road",
            "Road Type": "Dambar Road",
            "Floors": 4,
            "Plot Area SqFt": 22000.0,
            "Plinth Area SqFt": 14000.0,
            "Total Built Up Area SqFt": 38500.0,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 6,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Sewer System",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.561,

            "Longitude": 78.6594,
        },
        {
            "Status": "Approved",
            "Surveyor Name": "Ravi Chauhan",
            "Assessment Year": "2024-2025",
            "Date of Survey": datetime(2025, 3, 6, 9, 40),
            "Owner Name": "Kiran Patel",
            "Owner Father Name": "Ramesh Patel",
            "Mobile No": "9654321098",
            "Ward Name": "Ward 10 - Godown Area",
            "Is Slum": "no",
            "Parcel No": 178,
            "Unit No": 2,
            "old property Number": 4330,
            "Constructed Date": date(2005, 10, 23),
            "Respondent Name": "Kiran Patel",
            "Respondent Relationship": "Self",
            "City": "Etah",
            "Pincode": "207002",
            "House No": "Warehouse 3",
            "Locality": "Transport Nagar",
            "Colony": "Godown Cluster",
            "Tax Rate Zone": "RATE ZONE 3",
            "Property Ownership": "Individual (Single/Joint)",
            "Property Type": "Godown",
            "Property Uses": "Godown",
            "Situation": "Interior",
            "Road Type": "Rcc",
            "Floors": 1,
            "Plot Area SqFt": 6400.0,
            "Plinth Area SqFt": 6100.0,
            "Total Built Up Area SqFt": 6100.0,
            "Is Municipal Water Supply": "no",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Other",
            "Toilet Type": "Connected to Surface Drains",
            "Is Municipal Waste Service": "no",
            "Latitude": 27.5568,

            "Longitude": 78.6649,
        },
        {
            "Status": "Completed",
            "Surveyor Name": "Amit Verma",
            "Assessment Year": "2025-2026",
            "Date of Survey": datetime(2025, 12, 1, 11, 18),
            "Owner Name": "Meena Kumari",
            "Owner Father Name": "Shiv Prasad",
            "Mobile No": "9786543210",
            "Ward Name": "Ward 3 - Nehru Market",
            "Is Slum": "no",
            "Parcel No": 94,
            "Unit No": 1,
            "old property Number": 1999,
            "Constructed Date": date(2011, 4, 17),
            "Respondent Name": "Poonam Kumari",
            "Respondent Relationship": "Daughter",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "B-16",
            "Locality": "Nehru Market",
            "Colony": "Laxmi Nagar",
            "Tax Rate Zone": "RATE ZONE 2",
            "Property Ownership": "Individual (Single/Joint)",
            "Property Type": "Bakery",
            "Property Uses": "Bakery",
            "Situation": "Main Market",
            "Road Type": "Interlocking Road",
            "Floors": 2,
            "Plot Area SqFt": 540.25,
            "Plinth Area SqFt": 500.0,
            "Total Built Up Area SqFt": 910.6,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Septic Tank",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5595,

            "Longitude": 78.6621,
        },
        {
            "Status": "Draft",
            "Surveyor Name": "Neha Singh",
            "Assessment Year": "2026-2027",
            "Date of Survey": datetime(2026, 5, 19, 16, 48),
            "Owner Name": "Metro Retail LLP",
            "Owner Father Name": "N/A",
            "Mobile No": "9090901234",
            "Ward Name": "Ward 2 - Station Road",
            "Is Slum": "no",
            "Parcel No": 501,
            "Unit No": 8,
            "old property Number": 8080,
            "Constructed Date": date(2021, 1, 11),
            "Respondent Name": "Anita Gupta",
            "Respondent Relationship": "Other",
            "City": "Etah",
            "Pincode": "207001",
            "House No": "Showroom 2",
            "Locality": "Station Road",
            "Colony": "City Centre",
            "Tax Rate Zone": "RATE ZONE 1",
            "Property Ownership": "Limited Company",
            "Property Type": "Mall Showroom",
            "Property Uses": "Commercial",
            "Situation": "Main Market",
            "Road Type": "RCC Road",
            "Floors": 3,
            "Plot Area SqFt": 12500.0,
            "Plinth Area SqFt": 11000.0,
            "Total Built Up Area SqFt": 28640.0,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 5,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Sewer System",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5581,

            "Longitude": 78.6658,
        },
        {
            "Status": "Submitted",
            "Surveyor Name": "Imran Khan",
            "Assessment Year": "2024-2025",
            "Date of Survey": datetime(2025, 2, 21, 12, 12),
            "Owner Name": "Rajesh Patel",
            "Owner Father Name": "Kiran Patel",
            "Mobile No": "9234567810",
            "Ward Name": "Ward 7 - Subhash Nagar",
            "Is Slum": "no",
            "Parcel No": 276,
            "Unit No": 1,
            "old property Number": 3478,
            "Constructed Date": date(2009, 8, 3),
            "Respondent Name": "Sita Patel",
            "Respondent Relationship": "Mother",
            "City": "Etah",
            "Pincode": "207002",
            "House No": "31/4",
            "Locality": "Subhash Nagar",
            "Colony": "Patel Mohalla",
            "Tax Rate Zone": "RATE ZONE 3",
            "Property Ownership": "Joint",
            "Property Type": "Residential",
            "Property Uses": "Residential/Commercial",
            "Situation": "Interior",
            "Road Type": "Dambar Road",
            "Floors": 2,
            "Plot Area SqFt": 980.0,
            "Plinth Area SqFt": 860.4,
            "Total Built Up Area SqFt": 1540.2,
            "Is Municipal Water Supply": "yes",
            "Total Water Connection": 1,
            "Water Connection Id/Type": "Government Tap",
            "Toilet Type": "Other",
            "Is Municipal Waste Service": "yes",
            "Latitude": 27.5598,

            "Longitude": 78.6617,
        },
    ]


COLUMN_WIDTHS = {
    1: 8,
    2: 14,
    3: 18,
    4: 16,
    5: 32,
    6: 20,
    7: 28,
    8: 22,
    9: 15,
    10: 26,
    11: 12,
    12: 12,
    13: 12,
    14: 18,
    15: 18,
    16: 22,
    17: 22,
    18: 15,
    19: 15,
    20: 15,
    21: 18,
    22: 18,
    23: 16,
    24: 26,
    25: 32,
    26: 32,
    27: 16,
    28: 20,
    29: 12,
    30: 20,
    31: 20,
    32: 24,
    33: 24,
    34: 22,
    35: 28,
    36: 26,
    37: 24,
    38: 16,
    39: 16,
}

NUMBER_COLS = {1, 12, 13, 14, 29, 30, 31, 32, 34, 38, 39}
TEXT_COLS = {3, 7, 8, 9, 10, 16, 18, 19, 20, 21, 22, 5}
DATE_COLS = {6, 15}
CENTER_COLS = {2, 4, 11, 17, 23, 24, 25, 26, 27, 28, 33, 35, 36, 37}
LOCKED_COLS = {1, 5}

INPUT_MESSAGES = {
    1: ("SN", "Auto-generated serial number. Do not edit."),
    2: ("Status", "Select the current survey workflow status."),
    3: ("Surveyor Name", "Enter the name of the surveyor."),
    4: ("Assessment Year", "Select the financial assessment year."),
    5: ("Property Id", "Automatically generated from Parcel No, old property Number and Property Type."),
    6: ("Date of Survey", "Enter the survey date and time (dd-mm-yyyy hh:mm)."),
    9: ("Mobile No", "Enter exactly 10 digits."),
    11: ("Is Slum", "Select yes if the property is in a notified slum."),
    12: ("Parcel No", "Enter the parcel number."),
    13: ("Unit No", "Enter the unit / sub-property number."),
    14: ("old property Number", "Enter the legacy municipal property number."),
    17: ("Respondent Relationship", "Select the respondent's relationship to the owner."),
    19: ("Pincode", "Enter exactly 6 digits."),
    23: ("Tax Rate Zone", "Select the applicable municipal tax rate zone."),
    24: ("Property Ownership", "Select the legal ownership category."),
    25: ("Property Type", "Select the applicable property classification."),
    26: ("Property Uses", "Select the actual current use of the property."),
    29: ("Floors", "Enter the number of floors as a whole number (0 or more)."),
    30: ("Plot Area SqFt", "Enter plot area in square feet. Decimal values are allowed."),
    31: ("Plinth Area SqFt", "Enter plinth area in square feet. Decimal values are allowed."),
    32: ("Total Built Up Area SqFt", "Enter total built-up area in square feet."),
    33: ("Municipal Water", "Select whether the property has municipal water supply."),
    34: ("Water Connections", "Enter the total number of water connections (0 or more)."),
    35: ("Water Connection Type", "Select the water-connection or sanitation source type."),
    36: ("Toilet Type", "Select the sanitation / toilet arrangement."),
    37: ("Waste Service", "Select whether municipal waste collection is available."),
    38: ("Latitude", "Enter latitude in decimal degrees (-90 to 90). 6 decimal places."),
    39: ("Longitude", "Enter longitude in decimal degrees (-180 to 180). 6 decimal places."),
}

HEADER_COMMENTS = {
    5: "Formula: 800797-011-{Parcel No}-{old property Number}-{Property Type}. Blank source fields yield a blank ID instead of 800797-011---.",
    24: "Individual (Single/Joint) covers both sole and joint individual ownership per application business rules. Use Joint when the survey captured explicit joint ownership.",
    26: "Property Uses is the actual current use of the property (occupancy), which may differ from Property Type (classification).",
    35: "Permitted values include water sources (Government Tap, Borewell, Dug well) and some sanitation codes retained for legacy import compatibility. Prefer Toilet Type for sanitation.",
}


def build_validation_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Validation Lists")
    ws.sheet_properties.tabColor = "BDC3C7"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Named ranges on this sheet drive dropdowns on Survey Data. Do not delete values."
    ws.merge_cells("A1:O1")
    ws["A1"].font = font(italic=True, size=10, color=MUTED)
    ws.row_dimensions[1].height = 22

    for col_idx, (name, values) in enumerate(NAMED_LISTS.items(), start=1):
        cell = ws.cell(2, col_idx, name)
        cell.font = font(bold=True, color=WHITE, size=10)
        cell.fill = fill(PRIMARY)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        apply_border(cell)
        for row_idx, value in enumerate(values, start=3):
            vcell = ws.cell(row_idx, col_idx, value)
            vcell.font = font(size=11)
            vcell.alignment = Alignment(horizontal="left", vertical="center")
            apply_border(vcell)
            vcell.protection = Protection(locked=True)
        last_row = 2 + len(values)
        letter = col_letter(col_idx)
        # values start at row 3 (row 2 is the range name header)
        ref = f"'Validation Lists'!${letter}$3:${letter}${last_row}"
        defined = DefinedName(name, attr_text=ref)
        wb.defined_names.add(defined)
        ws.column_dimensions[letter].width = max(18, min(36, max(len(v) for v in values) + 2))

    ws.row_dimensions[2].height = 28
    ws.sheet_state = "hidden"
    ws.protection.sheet = True
    ws.protection.enable()


def build_banner(ws) -> None:
    last_col = col_letter(len(HEADERS))
    ws.merge_cells(f"A{TITLE_ROW}:{last_col}{TITLE_ROW}")
    title = ws["A1"]
    title.value = "Municipal Property Tax Survey  |  Data Entry Template"
    title.font = font(bold=True, color=WHITE, size=16)
    title.fill = fill(PRIMARY)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[TITLE_ROW].height = 32

    ws.merge_cells(f"A{META_ROW}:{last_col}{META_ROW}")
    meta = ws["A2"]
    meta.value = (
        "ULB 800797  |  Ward code 011  |  Property Id = 800797-011-{Parcel No}-{old property Number}-{Property Type}  |  "
        "SN and Property Id are formula-protected  |  Status is editable for field entry  |  v1.0"
    )
    meta.font = font(size=10, color=WHITE)
    meta.fill = fill(SECONDARY)
    meta.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[META_ROW].height = 22

    for start, end, label, color in GROUPS:
        start_l = col_letter(start)
        end_l = col_letter(end)
        ws.merge_cells(f"{start_l}{GROUP_ROW}:{end_l}{GROUP_ROW}")
        cell = ws.cell(GROUP_ROW, start)
        cell.value = label
        cell.font = font(bold=True, color=WHITE, size=10)
        cell.fill = fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start, end + 1):
            gcell = ws.cell(GROUP_ROW, col)
            gcell.fill = fill(color)
            apply_border(gcell, left_accent=col in GROUP_STARTS)
    ws.row_dimensions[GROUP_ROW].height = 20


def alignment_for(col_idx: int) -> Alignment:
    if col_idx in NUMBER_COLS:
        return Alignment(horizontal="right", vertical="center")
    if col_idx in DATE_COLS or col_idx in CENTER_COLS:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def number_format_for(col_idx: int) -> str:
    if col_idx == 6:
        return "DD-MM-YYYY HH:MM"
    if col_idx == 15:
        return "DD-MM-YYYY"
    if col_idx in {9, 19, 5}:
        return "@"
    if col_idx in {30, 31, 32}:
        return "#,##0.00"
    if col_idx in {38, 39}:
        return "0.000000"
    if col_idx in {1, 12, 13, 14, 29, 34}:
        return "0"
    return "General"


def sn_formula(row: int) -> str:
    # Skip SN (A) and Property Id (E) when deciding if the row is populated.
    return f'IF(COUNTA($B{row}:$D{row},$F{row}:$AM{row})>0,MAX($A$4:A{row - 1})+1,"")'


def property_id_formula(row: int) -> str:
    # L = Parcel No, N = old property Number, Y = Property Type
    return (
        f'IF(OR($L{row}="",$N{row}="",TRIM($Y{row})=""),"",'
        f'"800797-011-"&TEXT($L{row},"0")&"-"&TEXT($N{row},"0")&"-"&TRIM($Y{row}))'
    )


def build_survey_sheet(wb: Workbook, table_style_name: str) -> None:
    ws = wb.create_sheet("Survey Data", 0)
    ws.sheet_properties.tabColor = PRIMARY
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_view.view = "normal"
    ws.sheet_properties.outlinePr.summaryRight = True

    build_banner(ws)

    samples = sample_rows()

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(HEADER_ROW, col_idx, header)
        cell.style = "SurveyHeader"
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        apply_border(cell, left_accent=col_idx in GROUP_STARTS)
        if col_idx in HEADER_COMMENTS:
            cell.comment = comment(HEADER_COMMENTS[col_idx])
        ws.column_dimensions[col_letter(col_idx)].width = COLUMN_WIDTHS[col_idx]
    ws.row_dimensions[HEADER_ROW].height = 42

    for offset in range(TOTAL_DATA_ROWS):
        row = FIRST_DATA_ROW + offset
        zebra = WHITE if offset % 2 == 0 else BACKGROUND
        sample = samples[offset] if offset < len(samples) else None
        ws.row_dimensions[row].height = 22
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row, col_idx)
            cell.fill = fill(zebra)
            cell.font = font(size=11)
            cell.alignment = alignment_for(col_idx)
            cell.number_format = number_format_for(col_idx)
            apply_border(cell, left_accent=col_idx in GROUP_STARTS)
            locked = col_idx in LOCKED_COLS
            cell.protection = Protection(locked=locked)

            if col_idx == 1:
                cell.value = f"={sn_formula(row)}"
                continue
            if col_idx == 5:
                cell.value = f"={property_id_formula(row)}"
                continue
            if sample is not None:
                value = sample.get(header)
                if header in {"Mobile No", "Pincode"}:
                    cell.value = str(value)
                    cell.number_format = "@"
                else:
                    cell.value = value
                    if isinstance(value, datetime):
                        cell.number_format = "DD-MM-YYYY HH:MM"
                    elif isinstance(value, date):
                        cell.number_format = "DD-MM-YYYY"

    # Adjacent column groups cannot be outlined separately in Excel without
    # spacer columns. Visual grouping is the colored band on row 3 plus
    # accent left-borders at each group start.

    table = Table(displayName=TABLE_NAME, ref=table_ref())
    table.tableStyleInfo = TableStyleInfo(
        name=table_style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    add_validations(ws)
    add_status_conditional_formatting(ws)

    ws.freeze_panes = f"A{FIRST_DATA_ROW}"
    ws.auto_filter.ref = table_ref()
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"
    ws.print_area = table_ref()
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.25, footer=0.25)
    ws.oddHeader.left.text = "Municipal Property Tax Survey"
    ws.oddHeader.right.text = "Survey Data"
    ws.oddFooter.left.text = "SN & Property Id are formula-generated — do not overwrite"
    ws.oddFooter.right.text = "Page &P of &N"

    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True
    ws.protection.insertRows = True
    ws.protection.insertColumns = False
    ws.protection.deleteRows = False
    ws.protection.deleteColumns = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    ws.protection.formatCells = True
    ws.protection.formatColumns = True
    ws.protection.formatRows = True
    ws.protection.enable()


def add_validations(ws) -> None:
    dropdowns = {
        2: "StatusOptions",
        4: "AssessmentYearOptions",
        11: "SlumOptions",
        17: "RelationshipOptions",
        23: "TaxRateZoneOptions",
        24: "OwnershipOptions",
        25: "PropertyTypeOptions",
        26: "PropertyUseOptions",
        27: "SituationOptions",
        28: "RoadTypeOptions",
        33: "WaterSupplyOptions",
        35: "WaterConnectionTypeOptions",
        36: "ToiletTypeOptions",
        37: "WasteServiceOptions",
    }
    for col_idx, named_range in dropdowns.items():
        dv = DataValidation(
            type="list",
            formula1=named_range,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle="Invalid value",
            error=f"Select a value from the {named_range.replace('Options', '')} list.",
            errorStyle="stop",
        )
        title, prompt = INPUT_MESSAGES.get(col_idx, (HEADERS[col_idx - 1], "Select a valid value."))
        dv.promptTitle = title
        dv.prompt = prompt
        dv.add(data_range(col_idx))
        ws.add_data_validation(dv)

    # SN / Property Id input messages (still locked)
    for col_idx in (1, 5):
        dv = DataValidation(
            type="custom",
            formula1="TRUE",
            allow_blank=True,
            showErrorMessage=False,
            showInputMessage=True,
        )
        dv.promptTitle, dv.prompt = INPUT_MESSAGES[col_idx]
        dv.add(data_range(col_idx))
        ws.add_data_validation(dv)

    text_prompts = {
        3: INPUT_MESSAGES[3],
    }
    for col_idx, (title, prompt) in text_prompts.items():
        dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True, showInputMessage=True, showErrorMessage=False)
        dv.promptTitle, dv.prompt = title, prompt
        dv.add(data_range(col_idx))
        ws.add_data_validation(dv)

    mobile = DataValidation(
        type="custom",
        formula1=f'AND(LEN(TRIM(I{FIRST_DATA_ROW}))=10,ISNUMBER(--I{FIRST_DATA_ROW}),INT(--I{FIRST_DATA_ROW})=--I{FIRST_DATA_ROW})',
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid mobile number",
        error="Enter a valid 10-digit mobile number.",
        errorStyle="stop",
    )
    mobile.promptTitle, mobile.prompt = INPUT_MESSAGES[9]
    mobile.add(data_range(9))
    ws.add_data_validation(mobile)

    pincode = DataValidation(
        type="custom",
        formula1=f'AND(LEN(TRIM(S{FIRST_DATA_ROW}))=6,ISNUMBER(--S{FIRST_DATA_ROW}),INT(--S{FIRST_DATA_ROW})=--S{FIRST_DATA_ROW})',
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid pincode",
        error="Enter a valid 6-digit pincode.",
        errorStyle="stop",
    )
    pincode.promptTitle, pincode.prompt = INPUT_MESSAGES[19]
    pincode.add(data_range(19))
    ws.add_data_validation(pincode)

    for col_idx, header in ((12, "Parcel No"), (13, "Unit No"), (14, "old property Number")):
        letter = col_letter(col_idx)
        dv = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle=f"Invalid {header}",
            error=f"{header} must be a number (0 or greater).",
            errorStyle="stop",
        )
        dv.promptTitle, dv.prompt = INPUT_MESSAGES[col_idx]
        dv.add(data_range(col_idx))
        ws.add_data_validation(dv)

    date_survey = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="DATE(1990,1,1)",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid date",
        error="Enter a valid survey date and time.",
        errorStyle="stop",
    )
    date_survey.promptTitle, date_survey.prompt = INPUT_MESSAGES[6]
    date_survey.add(data_range(6))
    ws.add_data_validation(date_survey)

    constructed = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="DATE(1900,1,1)",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid date",
        error="Enter a valid constructed date.",
        errorStyle="stop",
    )
    constructed.promptTitle = "Constructed Date"
    constructed.prompt = "Enter the construction date of the property."
    constructed.add(data_range(15))
    ws.add_data_validation(constructed)

    floors = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid floors",
        error="Floors must be a whole number of 0 or more.",
        errorStyle="stop",
    )
    floors.promptTitle, floors.prompt = INPUT_MESSAGES[29]
    floors.add(data_range(29))
    ws.add_data_validation(floors)

    for col_idx in (30, 31, 32):
        dv = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle="Invalid area",
            error="Area must be a number of 0 or more. Decimals are allowed.",
            errorStyle="stop",
        )
        dv.promptTitle, dv.prompt = INPUT_MESSAGES[col_idx]
        dv.add(data_range(col_idx))
        ws.add_data_validation(dv)

    water_n = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid connection count",
        error="Total Water Connection must be a whole number of 0 or more.",
        errorStyle="stop",
    )
    water_n.promptTitle, water_n.prompt = INPUT_MESSAGES[34]
    water_n.add(data_range(34))
    ws.add_data_validation(water_n)

    # GPS: Latitude (-90 to 90) — col 38
    lat_dv = DataValidation(
        type="decimal",
        operator="between",
        formula1="-90",
        formula2="90",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid latitude",
        error="Latitude must be between -90 and 90.",
        errorStyle="stop",
    )
    lat_dv.promptTitle, lat_dv.prompt = INPUT_MESSAGES[38]
    lat_dv.add(data_range(38))
    ws.add_data_validation(lat_dv)

    # GPS: Longitude (-180 to 180) — col 39
    lng_dv = DataValidation(
        type="decimal",
        operator="between",
        formula1="-180",
        formula2="180",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid longitude",
        error="Longitude must be between -180 and 180.",
        errorStyle="stop",
    )
    lng_dv.promptTitle, lng_dv.prompt = INPUT_MESSAGES[39]
    lng_dv.add(data_range(39))
    ws.add_data_validation(lng_dv)


def add_status_conditional_formatting(ws) -> None:
    rng = f"A{FIRST_DATA_ROW}:{col_letter(len(HEADERS))}{LAST_DATA_ROW}"
    status_rng = data_range(2)
    for status, bg in STATUS_FILLS.items():
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'$B{FIRST_DATA_ROW}="{status}"'], fill=row_fill),
        )
        status_font = Font(bold=True, color=STATUS_FONTS[status], name=FONT_NAME, size=11)
        stronger = {
            "Completed": "C8EED8",
            "Draft": "FDE9A9",
            "Submitted": "CDE6F7",
            "Approved": "E4D2F0",
        }[status]
        ws.conditional_formatting.add(
            status_rng,
            FormulaRule(
                formula=[f'$B{FIRST_DATA_ROW}="{status}"'],
                fill=PatternFill(start_color=stronger, end_color=stronger, fill_type="solid"),
                font=status_font,
            ),
        )


def style_support_header(ws, last_col: str, title: str, subtitle: str, tab: str) -> None:
    ws.sheet_properties.tabColor = tab
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = font(bold=True, color=WHITE, size=16)
    ws["A1"].fill = fill(PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = font(size=10, color=WHITE)
    ws["A2"].fill = fill(SECONDARY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.5)


def build_progress_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Survey Progress")
    style_support_header(
        ws,
        "G",
        "Survey Progress Dashboard",
        "Completion is calculated from the 35 user-entered fields (SN and Property Id are excluded because they are formula-generated).",
        ACCENT,
    )
    headers = ["SN", "Property Id", "Status", "Filled Fields", "Total Fields", "Completion %", "Progress"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.font = font(bold=True, color=WHITE, size=11)
        cell.fill = fill(PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[4].height = 28

    # KPI cards on row 3
    last_progress = 4 + TOTAL_DATA_ROWS
    kpi_defs = [
        (1, f'="Populated"&CHAR(10)&COUNTIF(\'Survey Data\'!A{FIRST_DATA_ROW}:A{LAST_DATA_ROW},">0")'),
        (2, f'="Completed"&CHAR(10)&COUNTIF(\'Survey Data\'!B{FIRST_DATA_ROW}:B{LAST_DATA_ROW},"Completed")'),
        (3, f'="Draft"&CHAR(10)&COUNTIF(\'Survey Data\'!B{FIRST_DATA_ROW}:B{LAST_DATA_ROW},"Draft")'),
        (4, f'="Submitted"&CHAR(10)&COUNTIF(\'Survey Data\'!B{FIRST_DATA_ROW}:B{LAST_DATA_ROW},"Submitted")'),
        (5, f'="Approved"&CHAR(10)&COUNTIF(\'Survey Data\'!B{FIRST_DATA_ROW}:B{LAST_DATA_ROW},"Approved")'),
        (6, f'="Avg %"&CHAR(10)&TEXT(IFERROR(AVERAGEIF(F5:F{last_progress},">0"),0),"0%")'),
        (7, f'="Required"&CHAR(10)&{EDITABLE_FIELD_COUNT}'),
    ]
    for col, formula in kpi_defs:
        cell = ws.cell(3, col, formula)
        cell.font = font(bold=True, color=WHITE, size=10)
        cell.fill = fill(ACCENT if col in {1, 6, 7} else SECONDARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 36

    for offset in range(TOTAL_DATA_ROWS):
        src = FIRST_DATA_ROW + offset
        row = 5 + offset
        zebra = WHITE if offset % 2 == 0 else BACKGROUND
        filled = f'IF(\'Survey Data\'!A{src}="","",COUNTA(\'Survey Data\'!B{src}:D{src},\'Survey Data\'!F{src}:AM{src}))'
        formulas = [
            f'=IF(\'Survey Data\'!A{src}="","",\'Survey Data\'!A{src})',
            f'=IF(\'Survey Data\'!A{src}="","",\'Survey Data\'!E{src})',
            f'=IF(\'Survey Data\'!A{src}="","",\'Survey Data\'!B{src})',
            f"={filled}",
            f'=IF(\'Survey Data\'!A{src}="","",{EDITABLE_FIELD_COUNT})',
            f'=IF(OR(D{row}="",E{row}=""),"",D{row}/E{row})',
            f'=IF(F{row}="","",F{row})',
        ]
        for col, formula in enumerate(formulas, start=1):
            cell = ws.cell(row, col, formula)
            cell.fill = fill(zebra)
            cell.font = font(size=11)
            apply_border(cell)
            cell.protection = Protection(locked=True)
            if col in {1, 3, 4, 5}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0"
            elif col in {6, 7}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0%"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    bar = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=1,
        color=ACCENT,
        showValue=True,
        minLength=None,
        maxLength=None,
    )
    ws.conditional_formatting.add(f"G5:G{4 + TOTAL_DATA_ROWS}", bar)
    pct_bar = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=1,
        color=SECONDARY,
        showValue=True,
        minLength=None,
        maxLength=None,
    )
    ws.conditional_formatting.add(f"F5:F{4 + TOTAL_DATA_ROWS}", pct_bar)

    for col, width in enumerate([10, 34, 14, 16, 14, 16, 18], start=1):
        ws.column_dimensions[col_letter(col)].width = width
    ws.print_title_rows = "4:4"
    ws.print_area = f"A3:G{4 + TOTAL_DATA_ROWS}"
    ws.protection.sheet = True
    ws.protection.enable()


def build_actions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Actions")
    style_support_header(
        ws,
        "F",
        "Survey Actions (documentation + optional tracker)",
        "Excel cannot host real web buttons without VBA. The Next.js admin app should render Demand Notice, Print, View, Edit and Delete as UI buttons. This sheet is an auxiliary tracker only — it does not change the 37-column Survey Data schema.",
        SECONDARY,
    )
    notes = [
        "Demand Notice — generate / print the municipal demand notice for the selected property (web app).",
        "Print — print the survey summary or notice pack (web app).",
        "View — open the read-only survey record (web app).",
        "Edit — open the survey editor with Clerk-authenticated permissions (web app).",
        "Delete — soft-delete the survey in Postgres (web app; never delete rows here to stay import-safe).",
    ]
    ws["A3"] = "Application actions (implement in Next.js, not in this workbook)"
    ws["A3"].font = font(bold=True, size=12, color=PRIMARY)
    ws.merge_cells("A3:F3")
    for i, line in enumerate(notes, start=4):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.cell(i, 1, line).font = font(size=11)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 18

    headers = ["SN", "Property Id", "Status", "Suggested Action", "Operator Notes", "Implemented in"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(10, col, header)
        cell.font = font(bold=True, color=WHITE, size=11)
        cell.fill = fill(PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[10].height = 28

    for offset in range(TOTAL_DATA_ROWS):
        src = FIRST_DATA_ROW + offset
        row = 11 + offset
        zebra = WHITE if offset % 2 == 0 else BACKGROUND
        values = [
            f'=IF(\'Survey Data\'!A{src}="","",\'Survey Data\'!A{src})',
            f'=IF(\'Survey Data\'!A{src}="","",\'Survey Data\'!E{src})',
            f'=IF(\'Survey Data\'!A{src}="","",\'Survey Data\'!B{src})',
            None,
            "",
            f'=IF(\'Survey Data\'!A{src}="","","Next.js UI")',
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.fill = fill(zebra)
            cell.font = font(size=11)
            apply_border(cell)
            cell.alignment = Alignment(horizontal="center" if col != 5 else "left", vertical="center")
            cell.protection = Protection(locked=col not in {4, 5})
        ws.row_dimensions[row].height = 20

    action_dv = DataValidation(
        type="list",
        formula1="ActionOptions",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
        error="Select Demand Notice, Print, View, Edit or Delete.",
        errorTitle="Invalid action",
    )
    action_dv.promptTitle = "Suggested Action"
    action_dv.prompt = "Optional tracker only. Real actions belong in the Next.js application."
    action_dv.add(f"D11:D{10 + TOTAL_DATA_ROWS}")
    ws.add_data_validation(action_dv)

    # Seed a few suggested actions on sample rows
    seeds = ["View", "Edit", "Demand Notice", "Print", "View", "Edit", "View", "Demand Notice", "View", "Edit"]
    for i, action in enumerate(seeds):
        ws.cell(11 + i, 4, action)

    for col, width in enumerate([10, 34, 14, 18, 36, 16], start=1):
        ws.column_dimensions[col_letter(col)].width = width
    ws.print_title_rows = "10:10"
    ws.print_area = f"A3:F{10 + TOTAL_DATA_ROWS}"
    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.enable()


def build_field_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("Field Guide")
    style_support_header(
        ws,
        "I",
        "Field Guide  |  Excel to NestJS DTO mapping",
        "Stable 37-column contract for import/export. DTO names are camelCase suggestions for apps/api survey DTOs. Aadhaar is documented as a future optional field and is intentionally absent from the primary sheet.",
        "1F618D",
    )
    guide_headers = [
        "#",
        "Excel Column",
        "Data Type",
        "Required",
        "Validation",
        "Description",
        "Example",
        "NestJS DTO Field",
        "DTO Type",
    ]
    rows = [
        (1, "SN", "number", "Auto", "Formula, protected", "Serial number for populated rows", "1", "serialNumber", "number"),
        (2, "Status", "enum", "Yes", "Dropdown", "Workflow status for field/QC entry", "Completed", "status", "string"),
        (3, "Surveyor Name", "text", "Yes", "Text", "Name of the enumerator", "Amit Verma", "surveyorName", "string"),
        (4, "Assessment Year", "enum", "Yes", "Dropdown", "Financial assessment year", "2025-2026", "assessmentYear", "string"),
        (5, "Property Id", "text", "Auto", "Formula, protected", "800797-011-{parcel}-{oldNo}-{type}", "800797-011-125-4587-Residential", "propertyId", "string"),
        (6, "Date of Survey", "datetime", "Yes", "Valid date", "Capture / submit timestamp", "12-06-2025 10:35", "dateOfSurvey", "datetime"),
        (7, "Owner Name", "text", "Yes", "Text", "Primary owner or entity name", "Ramesh Kumar Sharma", "ownerName", "string"),
        (8, "Owner Father Name", "text", "No", "Text", "Father / husband name; N/A for organisations", "Ram Lal Sharma", "ownerFatherName", "string"),
        (9, "Mobile No", "text", "Yes", "Exactly 10 digits", "Stored as text to preserve digits", "9876501234", "mobileNo", "string"),
        (10, "Ward Name", "text", "Yes", "Text", "Ward display name", "Ward 11 - Gandhi Nagar", "wardName", "string"),
        (11, "Is Slum", "enum", "Yes", "yes / no", "Notified slum flag", "no", "isSlum", "string"),
        (12, "Parcel No", "number", "Yes", "Numeric >= 0", "Parcel number used in Property Id", "125", "parcelNo", "number"),
        (13, "Unit No", "number", "Yes", "Numeric >= 0", "Unit / sub-number", "1", "unitNo", "number"),
        (14, "old property Number", "number", "Yes", "Numeric >= 0", "Legacy municipal property number", "4587", "oldPropertyNumber", "number"),
        (15, "Constructed Date", "date", "No", "Valid date", "Construction date", "14-03-2008", "constructedDate", "date"),
        (16, "Respondent Name", "text", "No", "Text", "Person interviewed on site", "Ramesh Kumar Sharma", "respondentName", "string"),
        (17, "Respondent Relationship", "enum", "No", "Dropdown", "Relation to owner", "Self", "respondentRelationship", "string"),
        (18, "City", "text", "Yes", "Text", "City / ULB name", "Etah", "city", "string"),
        (19, "Pincode", "text", "Yes", "Exactly 6 digits", "Indian PIN code as text", "207001", "pincode", "string"),
        (20, "House No", "text", "No", "Text", "Door / house / plot identifier", "12-A", "houseNo", "string"),
        (21, "Locality", "text", "No", "Text", "Locality / mohalla", "Gandhi Nagar", "locality", "string"),
        (22, "Colony", "text", "No", "Text", "Colony name", "Shastri Colony", "colony", "string"),
        (23, "Tax Rate Zone", "enum", "Yes", "RATE ZONE 1–4", "Municipal rate zone used by tax engine", "RATE ZONE 1", "taxRateZone", "string"),
        (24, "Property Ownership", "enum", "Yes", "Dropdown", "Legal ownership category", "Individual (Single/Joint)", "propertyOwnership", "string"),
        (25, "Property Type", "enum", "Yes", "Dropdown", "Classification; also feeds Property Id", "Residential", "propertyType", "string"),
        (26, "Property Uses", "enum", "Yes", "Dropdown", "Actual current use / occupancy", "Residential Self", "propertyUses", "string"),
        (27, "Situation", "enum", "Yes", "Dropdown", "Location relative to market/road/slum", "Main Road", "situation", "string"),
        (28, "Road Type", "enum", "Yes", "Dropdown", "Access road surface", "Dambar Road", "roadType", "string"),
        (29, "Floors", "integer", "Yes", "Whole number >= 0", "Storey count; detailed floors live on Floor Details", "2", "floors", "number"),
        (30, "Plot Area SqFt", "decimal", "Yes", "Number >= 0", "Plot area in square feet", "1200.50", "plotAreaSqFt", "number"),
        (31, "Plinth Area SqFt", "decimal", "No", "Number >= 0", "Plinth area in square feet", "980.25", "plinthAreaSqFt", "number"),
        (32, "Total Built Up Area SqFt", "decimal", "Yes", "Number >= 0", "Total built-up area in square feet", "1760.75", "totalBuiltUpAreaSqFt", "number"),
        (33, "Is Municipal Water Supply", "enum", "Yes", "yes / no", "Municipal water availability", "yes", "isMunicipalWaterSupply", "string"),
        (34, "Total Water Connection", "integer", "No", "Whole number >= 0", "Count of water connections", "1", "totalWaterConnection", "number"),
        (35, "Water Connection Id/Type", "enum", "No", "Dropdown", "Source / connection type (legacy mixed list)", "Government Tap", "waterConnectionIdType", "string"),
        (36, "Toilet Type", "enum", "No", "Dropdown", "Sanitation arrangement", "Sewer System", "toiletType", "string"),
        (37, "Is Municipal Waste Service", "enum", "Yes", "yes / no", "Door-to-door / municipal waste flag", "yes", "isMunicipalWasteService", "string"),
        (38, "Latitude", "decimal", "No", "Decimal -90 to 90, 6 dp", "GPS latitude captured by surveyor device", "27.558300", "latitude", "number"),
        (39, "Longitude", "decimal", "No", "Decimal -180 to 180, 6 dp", "GPS longitude captured by surveyor device", "78.662800", "longitude", "number"),
        (
            "—",
            "Aadhaar Number (future / optional)",
            "text",
            "No",
            "12 digits if collected later",
            "NOT in the 37-column primary schema. If the application adds Aadhaar, validate 12 digits, store as text, never publish in sample workbooks, and keep it off exports by default.",
            "XXXXXXXX1234 (masked)",
            "aadhaarNumber",
            "string | null",
        ),
    ]
    for col, header in enumerate(guide_headers, start=1):
        cell = ws.cell(4, col, header)
        cell.font = font(bold=True, color=WHITE, size=11)
        cell.fill = fill(PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[4].height = 28

    for r_idx, row in enumerate(rows, start=5):
        zebra = WHITE if (r_idx - 5) % 2 == 0 else BACKGROUND
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.fill = fill("FDEBD0") if r_idx == 5 + 37 else fill(zebra)
            cell.font = font(size=10, italic=r_idx == 5 + 37)
            cell.alignment = Alignment(
                horizontal="center" if c_idx in {1, 3, 4, 8, 9} else "left",
                vertical="center",
                wrap_text=True,
            )
            apply_border(cell)
        ws.row_dimensions[r_idx].height = 32 if r_idx == 5 + 37 else 22

    widths = [6, 32, 14, 12, 28, 48, 34, 26, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[col_letter(i)].width = width
    ws.auto_filter.ref = f"A4:I{4 + len(rows)}"
    ws.print_title_rows = "4:4"
    ws.print_area = f"A1:I{4 + len(rows)}"
    ws.page_setup.orientation = "landscape"


def build_floor_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Floor Details")
    style_support_header(
        ws,
        "H",
        "Reference | Floor Details",
        "The primary Floors column is a storey count only. Detailed floor-wise area, usage and construction stay here so the 37-column import schema remains stable. Conversion: 1 SqMt = 10.7639104167 SqFt.",
        "148F77",
    )
    headers = [
        "SN",
        "Property Id",
        "Floor Position",
        "Area SqFt",
        "Area SqMt",
        "Usage Type",
        "Usage Factor",
        "Construction Type",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.font = font(bold=True, color=WHITE, size=11)
        cell.fill = fill(PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[4].height = 28

    example = [
        [1, None, "Ground Floor", 625, None, "Residential", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [1, None, "First Floor", 540, None, "Residential", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [2, None, "Ground Floor", 820.5, None, "Commercial", "Rented", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [3, None, "Ground Floor", 1100, None, "Residential", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [3, None, "First Floor", 900, None, "Commercial", "Rented", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [3, None, "Second Floor", 650.4, None, "Residential", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [11, None, "Ground Floor", 14000, None, "Hospital Nursing Pathology", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [11, None, "First Floor", 12000, None, "Hospital Nursing Pathology", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [11, None, "Second Floor", 6500, None, "Hospital Nursing Pathology", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
        [11, None, "Third Floor", 6000, None, "Hospital Nursing Pathology", "Self Occupied", "Pakka Building with R.C.C Roof or R.B. Roof"],
    ]
    for r_idx, row in enumerate(example, start=5):
        zebra = WHITE if (r_idx - 5) % 2 == 0 else BACKGROUND
        sn, _, floor, sqft, _, usage, factor, construction = row
        src = FIRST_DATA_ROW + int(sn) - 1
        values = [
            sn,
            f"='Survey Data'!E{src}",
            floor,
            sqft,
            f"=ROUND(D{r_idx}/10.7639104167,4)",
            usage,
            factor,
            construction,
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.fill = fill(zebra)
            cell.font = font(size=11)
            apply_border(cell)
            if c_idx in {4, 5}:
                cell.number_format = "0.00" if c_idx == 4 else "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in {1}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r_idx].height = 22

    ws.merge_cells("A16:H16")
    ws["A16"] = (
        "Legacy narrative example: Ground Floor - 625 SqFt - 58.0644 SqMt || Usage Type - Residential || "
        "Usage Factor - Self Occupied || Construction Type - Pakka Building with R.C.C Roof or R.B. Roof. "
        "Do not paste this narrative into the Floors column on Survey Data."
    )
    ws["A16"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A16"].font = font(italic=True, size=10, color=MUTED)
    ws.row_dimensions[16].height = 36

    for col, width in enumerate([8, 34, 18, 14, 14, 28, 18, 48], start=1):
        ws.column_dimensions[col_letter(col)].width = width
    ws.print_title_rows = "4:4"
    ws.print_area = "A1:H16"


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    style_support_header(
        ws,
        "B",
        "README | How to use this workbook",
        "Data-entry / import template for the NestJS + Next.js municipal property survey application. No macros. No password.",
        PRIMARY,
    )
    sections = [
        ("How to enter data",
         "Use the Survey Data sheet only for the 39-column record (37 base + Longitude + Latitude). Start from the first empty row after the sample data. Dropdowns appear when you select a cell. Extra blank rows (65) are pre-created so you can type without inserting rows."),
        ("Which cells are protected",
         "SN (column A) and Property Id (column E) are locked because they are formulas. The sheet is protected with NO password — developers can Review → Unprotect Sheet. Status is NOT locked; it is a manual workflow dropdown for field entry. If the web application later owns Status, lock that column in the exporter."),
        ("How Property Id is generated",
         "Property Id = 800797-011-{Parcel No}-{old property Number}-{Property Type}. Example: 800797-011-125-4587-Residential. If Parcel No, old property Number, or Property Type is blank, the cell stays empty instead of emitting 800797-011---."),
        ("How SN is generated",
         "SN increments only for rows that have user-entered values. Completely empty template rows stay blank."),
        ("How dropdowns work",
         "Lists live on the hidden Validation Lists sheet and are exposed as named ranges (StatusOptions, PropertyTypeOptions, …). Do not type free-text into dropdown columns if you need a clean import."),
        ("How completion percentage works",
         "Survey Progress counts non-empty values among the 37 user-entered fields (39 total minus SN and Property Id). Completion % = Filled Fields / 37. Data bars visualise the ratio."),
        ("Which fields are controlled by the application",
         "In production, Property Id is also derived by @workspace/validation formatPropertyId (ULB-Ward-Parcel-Unit-Use). This template uses the requested 800797-011-parcel-oldNo-type pattern for Excel data entry. NestJS should treat Excel Property Id as a source value and re-derive on import when ULB/ward codes are known."),
        ("Mapping to NestJS / Next.js",
         "Keep column names and order unchanged. See Field Guide for Excel Column → camelCase DTO → type. The worker import path should map these headers onto the existing survey DTO. Do not add decorative symbols inside data values. Dates are real Excel datetimes."),
        ("Actions and progress",
         "Demand Notice / Print / View / Edit / Delete are documented on the Actions sheet. Implement them as Next.js buttons. Do not add VBA."),
        ("GPS columns (Longitude / Latitude)",
         "Columns 38–39 store the device GPS fix. Longitude is -180 to 180, Latitude is -90 to 90, both with 6 decimal places (Decimal(9,6) in Prisma). The surveyor mobile app auto-fills these; manual entry is allowed for corrections. Leave blank if GPS was unavailable."),
        ("Aadhaar",
         "Aadhaar Number is intentionally omitted from the 39-column schema. The Field Guide describes future 12-digit validation if the product adds it. Sample data contains no Aadhaar values."),
        ("Printing",
         "Survey Data prints landscape, fitted to 1 page wide, with the header row repeated. Hide unused columns before printing a ward extract if needed."),
        ("Visual grouping",
         "The 39-column order is fixed. Groups are shown as the colored band above the table (Survey / Workflow, Owner / Respondent, Address, Classification, Area, Utility, GPS / Location) with an accent border at each group start. Excel cannot outline adjacent groups without inserting spacer columns, so outline grouping is not used."),
        ("Unavoidable Excel limitations",
         "Excel is not a responsive web UI; wide surveys still need horizontal scrolling. Table auto-expansion of protection/validation on brand-new rows below the pre-created range may require Unprotect Sheet. Conditional formatting cannot draw a physical left-edge bar independently of cell fill, so status colour is a full-row tint plus a stronger Status cell. Named-range dropdowns on a hidden sheet work in Microsoft Excel; some third-party spreadsheet apps may ignore them."),
    ]
    ws["A4"] = "Topic"
    ws["B4"] = "Guidance"
    for col in (1, 2):
        ws.cell(4, col).font = font(bold=True, color=WHITE, size=11)
        ws.cell(4, col).fill = fill(PRIMARY)
        ws.cell(4, col).alignment = Alignment(horizontal="center", vertical="center")
        apply_border(ws.cell(4, col))
    for i, (title, body) in enumerate(sections, start=5):
        zebra = WHITE if (i - 5) % 2 == 0 else BACKGROUND
        a = ws.cell(i, 1, title)
        b = ws.cell(i, 2, body)
        for cell in (a, b):
            cell.fill = fill(zebra)
            cell.font = font(size=11, bold=cell.column == 1)
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
            apply_border(cell)
        ws.row_dimensions[i].height = 64
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A5"
    ws.print_title_rows = "4:4"
    ws.print_area = f"A1:B{4 + len(sections)}"
    ws.page_setup.orientation = "landscape"


def configure_workbook(wb: Workbook) -> None:
    wb.properties.title = "Municipal Property Tax Survey Data Template"
    wb.properties.creator = "SDV EduTech Survey Platform"
    wb.properties.subject = "survey_data_modern data-entry / import template"
    wb.properties.description = (
        "37-column municipal property survey template with validations, formulas, and NestJS DTO mapping."
    )
    wb.properties.keywords = "survey, municipal, property tax, nestjs, nextjs"
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True


def create_workbook() -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    configure_workbook(wb)
    register_named_style(wb)
    table_style = register_custom_table_style(wb)
    build_validation_sheet(wb)
    build_survey_sheet(wb, table_style)
    build_progress_sheet(wb)
    build_actions_sheet(wb)
    build_field_guide(wb)
    build_floor_sheet(wb)
    build_readme(wb)
    # Keep Survey Data first
    wb.move_sheet("Survey Data", offset=-len(wb.sheetnames) + 1)
    OUTPUT.unlink(missing_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


def expected_property_id(parcel: object, old_no: object, prop_type: object) -> str:
    if parcel in ("", None) or old_no in ("", None) or not str(prop_type or "").strip():
        return ""
    return f"800797-011-{int(float(parcel))}-{int(float(old_no))}-{str(prop_type).strip()}"


def verify(path: Path) -> list[str]:
    issues: list[str] = []
    wb = load_workbook(path)
    names = wb.sheetnames
    if "Survey Data" not in names:
        issues.append("Missing sheet Survey Data")
        return issues
    ws = wb["Survey Data"]
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, 40)]
    headers = [h for h in headers if h]
    if headers != HEADERS:
        issues.append(f"Header mismatch: {headers}")
    if ws.cell(HEADER_ROW, len(HEADERS) + 1).value not in (None, ""):
        issues.append(f"Unexpected extra column {len(HEADERS) + 1} in header row")
    named = set(wb.defined_names.keys())
    for expected in NAMED_LISTS:
        if expected not in named:
            issues.append(f"Missing named range {expected}")
    if "Validation Lists" not in names:
        issues.append("Missing Validation Lists sheet")
    else:
        if wb["Validation Lists"].sheet_state != "hidden":
            issues.append("Validation Lists is not hidden")
    sn = ws["A5"].value
    pid = ws["E5"].value
    if not (isinstance(sn, str) and sn.startswith("=") and "MAX(" in sn):
        issues.append(f"SN formula missing/invalid: {sn}")
    if not (isinstance(pid, str) and pid.startswith("=") and "800797-011-" in pid):
        issues.append(f"Property Id formula missing/invalid: {pid}")
    blank_pid = ws[f"E{FIRST_DATA_ROW + SAMPLE_COUNT}"].value
    if not (isinstance(blank_pid, str) and "OR(" in blank_pid):
        issues.append("Blank-row Property Id formula does not guard empty parts")
    sample = sample_rows()[0]
    expected_id = expected_property_id(sample["Parcel No"], sample["old property Number"], sample["Property Type"])
    if expected_id != "800797-011-125-4587-Residential":
        issues.append(f"Unexpected sample id logic: {expected_id}")
    # empty ID path
    if expected_property_id("", 1, "Residential") != "":
        issues.append("Blank parcel should not produce an ID")
    dvs = ws.data_validations.dataValidation
    if len(dvs) < 10:
        issues.append(f"Too few data validations: {len(dvs)}")
    cf_rules = [rule for rules in ws.conditional_formatting._cf_rules.values() for rule in rules]
    statuses_cf = {str(rule.formula[0]) for rule in cf_rules if rule.formula}
    for status in ("Completed", "Draft", "Submitted", "Approved"):
        if not any(f'$B{FIRST_DATA_ROW}="{status}"' in item for item in statuses_cf):
            issues.append(f"Missing conditional formatting for {status}")
    if ws.freeze_panes != f"A{FIRST_DATA_ROW}":
        issues.append(f"Freeze panes unexpected: {ws.freeze_panes}")
    if not ws.auto_filter.ref:
        issues.append("AutoFilter missing")
    if not ws.print_area:
        issues.append("Print area missing")
    elif table_ref().replace("$", "") not in ws.print_area.replace("$", ""):
        issues.append(f"Print area unexpected: {ws.print_area}")
    titles = (ws.print_title_rows or "").replace("$", "")
    if titles != f"{HEADER_ROW}:{HEADER_ROW}":
        issues.append(f"Print titles unexpected: {ws.print_title_rows}")
    if not ws.protection.sheet:
        issues.append("Sheet protection not enabled")
    if ws["A5"].protection.locked is not True:
        issues.append("SN cell is not locked")
    if ws["E5"].protection.locked is not True:
        issues.append("Property Id cell is not locked")
    if ws["B5"].protection.locked is not False:
        issues.append("Status should remain editable")
    tables = list(ws.tables.values())
    if not tables:
        issues.append("Excel table missing")
    else:
        if tables[0].ref != table_ref():
            issues.append(f"Table ref {tables[0].ref} != {table_ref()}")
        if len(tables[0].tableColumns) != len(HEADERS):
            issues.append(f"Table column count {len(tables[0].tableColumns)} != {len(HEADERS)}")
    populated = 0
    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + SAMPLE_COUNT):
        if ws.cell(row, 7).value:
            populated += 1
        mobile = str(ws.cell(row, 9).value)
        pin = str(ws.cell(row, 19).value)
        if len(mobile) != 10 or not mobile.isdigit():
            issues.append(f"Row {row} mobile invalid: {mobile}")
        if len(pin) != 6 or not pin.isdigit():
            issues.append(f"Row {row} pincode invalid: {pin}")
        status = ws.cell(row, 2).value
        if status not in NAMED_LISTS["StatusOptions"]:
            issues.append(f"Row {row} status invalid: {status}")
        for col in range(1, 38):
            val = ws.cell(row, col).value
            if isinstance(val, str) and val.startswith("#"):
                issues.append(f"Formula error-looking value at {col_letter(col)}{row}: {val}")
    if populated != SAMPLE_COUNT:
        issues.append(f"Expected {SAMPLE_COUNT} sample owner names, found {populated}")
    required_sheets = {"Survey Data", "Validation Lists", "Field Guide", "Survey Progress", "README", "Actions", "Floor Details"}
    missing = required_sheets - set(names)
    if missing:
        issues.append(f"Missing sheets: {missing}")
    if wb.vba_archive:
        issues.append("Workbook unexpectedly contains VBA")
    return issues


def main() -> None:
    path = create_workbook()
    issues = verify(path)
    print(f"Wrote {path}")
    print(f"Size {path.stat().st_size} bytes")
    if issues:
        print("VERIFICATION ISSUES:")
        for issue in issues:
            print(f" - {issue}")
        raise SystemExit(1)
    print("Verification passed.")
    print(f"Primary columns: {len(HEADERS)}")
    print(f"Sample rows: {SAMPLE_COUNT}")
    print(f"Template rows: {TOTAL_DATA_ROWS}")


if __name__ == "__main__":
    main()
