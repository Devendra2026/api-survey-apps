#!/usr/bin/env python3
"""Rebuild Etah ward survey Excel floors to match WARD 1.xlsx QC layout."""

from __future__ import annotations

import argparse
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Source premium columns (1-based)
SRC_SURVEY_ID = 2
SRC_OWNER = 17
SRC_FATHER = 18
SRC_MOBILE = 19
SRC_WARD = 8
SRC_PARCEL = 10
SRC_UNIT = 11
SRC_CITY = 25
SRC_PIN = 26
SRC_HOUSE = 22
SRC_COLONY = 24
SRC_TAX_ZONE = 32
SRC_PROP_TYPE = 28
SRC_PROP_USE = 29
SRC_ROAD = 31
SRC_PLOT = 33
SRC_PLINTH = 34
SRC_BUILT = 35

# (name, area_col, usage_factor_col, usage_type_col)
SRC_FLOOR_BLOCKS = [
    ("Basement", 36, 37, 38),
    ("Ground Floor", 40, 41, 42),
    ("First Floor", 44, 45, 46),
    ("Second Floor", 48, 49, 50),
    ("Third Floor", 52, 53, 54),
    ("Fourth Floor", 56, 57, 58),
    ("Fifth Floor", 60, 61, 62),
    ("Sixth Floor", 64, 65, 66),
    ("Open Land", 68, 69, 70),
]

FLOOR_GROUPS = [
    "Basement",
    "Ground Floor",
    "First Floor",
    "Second Floor",
    "Third Floor",
    "Fourth Floor",
    "Fifth Floor",
    "Sixth Floor",
    "Seventh Floor",
]

UPPER_FLOOR_INDEX = {
    "First Floor": 1,
    "Second Floor": 2,
    "Third Floor": 3,
    "Fourth Floor": 4,
    "Fifth Floor": 5,
    "Sixth Floor": 6,
    "Seventh Floor": 7,
}

IDENTITY_HEADERS = [
    "SN",
    "Survey Id",
    "Owner Name",
    "Owner Father Name",
    "Mobile No",
    "Ward Name",
    "Parcel No",
    "Unit Number",
    "City",
    "Pincode",
    "House No",
    "Colony",
    "Tax Rate Zone",
    "Property Type",
    "Property Use",
    "Road Type",
]

NON_RES_KEYWORDS = (
    "commercial",
    "godown",
    "shop",
    "non-residential",
    "non residential",
    "nonresidential",
    "mix",
)
RES_KEYWORDS = ("residential", "self", "rented", "residence")
NA_VALUES = {"", "n/a", "na"}

HEADER_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

OUT_FLOORS = 17
OUT_MATRIX_START = 18  # Basement Residential
OUT_OPEN = 36
OUT_PLOT = 38
OUT_PLINTH = 39
OUT_BUILT = 40
OUT_LAST = 40


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_na(value: Any) -> bool:
    return text(value).lower() in NA_VALUES


def parse_area(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:
            return None
        return float(value)
    s = text(value)
    if not s or s.lower() in NA_VALUES:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def area_or_zero(value: Any) -> float:
    parsed = parse_area(value)
    return 0.0 if parsed is None else parsed


def is_residential(usage_factor: Any, usage_type: Any) -> bool:
    factor = text(usage_factor).lower()
    utype = text(usage_type).lower()
    if factor and any(kw in factor for kw in NON_RES_KEYWORDS):
        return False
    if any(kw in factor for kw in RES_KEYWORDS) or any(kw in utype for kw in RES_KEYWORDS):
        return True
    if (not factor or factor in NA_VALUES) and (not utype or utype in NA_VALUES):
        return True
    return False


def parcel_sort_key(value: Any) -> tuple[int, float, str]:
    s = text(value)
    if not s or s.upper() in ("N/A", "NA"):
        return (1, 0.0, "")
    if re.fullmatch(r"\d+", s):
        return (0, float(int(s)), s)
    try:
        return (0, float(s), s)
    except ValueError:
        return (0, float("inf"), s)


def compute_floors_code(active: dict[str, float]) -> str:
    """WARD 1 pattern: optional B + optional G + F{max upper} else P."""
    has_b = active.get("Basement", 0) > 0
    has_g = active.get("Ground Floor", 0) > 0
    upper = [idx for name, idx in UPPER_FLOOR_INDEX.items() if active.get(name, 0) > 0]
    parts: list[str] = []
    if has_b:
        parts.append("B")
    if has_g:
        parts.append("G")
    if upper:
        parts.append(f"F{max(upper)}")
    return "".join(parts) if parts else "P"


def style_header_cell(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def resolve_input(path: Path, zip_entry: str) -> Path:
    if path.suffix.lower() == ".zip":
        dest = path.parent / "_etah_extract"
        dest.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(path) as zf:
            zf.extract(zip_entry, dest)
        return dest / zip_entry
    if not path.exists():
        raise SystemExit(f"Input not found: {path}")
    return path


def load_source_rows(ws) -> list[list[Any]]:
    rows: list[list[Any]] = []
    max_col = max(71, ws.max_column)
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or text(v) == "" for v in row):
            continue
        rows.append(row)
    return rows


def build_header(ws) -> None:
    # Identity cols 1–16
    for c, title in enumerate(IDENTITY_HEADERS, start=1):
        cell = ws.cell(1, c, title)
        style_header_cell(cell)
        for r in range(2, 5):
            style_header_cell(ws.cell(r, c))
        ws.merge_cells(start_row=1, start_column=c, end_row=4, end_column=c)

    # Floors band Q1:AK1 (17–37)
    ws.cell(1, OUT_FLOORS, "Floors")
    for c in range(OUT_FLOORS, OUT_OPEN + 2):
        style_header_cell(ws.cell(1, c))
    ws.merge_cells(start_row=1, start_column=OUT_FLOORS, end_row=1, end_column=OUT_OPEN + 1)

    for r in range(2, 4):
        style_header_cell(ws.cell(r, OUT_FLOORS))
    floor_label = ws.cell(4, OUT_FLOORS, "Floor")
    style_header_cell(floor_label)

    col = OUT_MATRIX_START
    for group in FLOOR_GROUPS:
        ws.cell(2, col, group)
        for c in (col, col + 1):
            style_header_cell(ws.cell(2, c))
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)

        ws.cell(3, col, "Residential")
        ws.cell(3, col + 1, "Non-Residential")
        for c in (col, col + 1):
            style_header_cell(ws.cell(3, c))
            style_header_cell(ws.cell(4, c))
            ws.merge_cells(start_row=3, start_column=c, end_row=4, end_column=c)
        col += 2

    # Open Land (36–37)
    ws.cell(2, OUT_OPEN, "Open Land (Plot)")
    for c in (OUT_OPEN, OUT_OPEN + 1):
        style_header_cell(ws.cell(2, c))
    ws.merge_cells(start_row=2, start_column=OUT_OPEN, end_row=2, end_column=OUT_OPEN + 1)

    ws.cell(3, OUT_OPEN, "Open Land")
    for c in (OUT_OPEN, OUT_OPEN + 1):
        style_header_cell(ws.cell(3, c))
        style_header_cell(ws.cell(4, c))
    ws.merge_cells(start_row=3, start_column=OUT_OPEN, end_row=4, end_column=OUT_OPEN + 1)

    for c, title in (
        (OUT_PLOT, "Plot Area SqFt"),
        (OUT_PLINTH, "Plinth Area SqFt"),
        (OUT_BUILT, "Total Built Up Area SqFt"),
    ):
        cell = ws.cell(1, c, title)
        style_header_cell(cell)
        for r in range(2, 5):
            style_header_cell(ws.cell(r, c))
        ws.merge_cells(start_row=1, start_column=c, end_row=4, end_column=c)

    for r in range(1, 5):
        ws.row_dimensions[r].height = 22


def map_identity(src: list[Any], sno: int) -> list[Any]:
    def cell(idx: int) -> Any:
        return src[idx - 1] if idx - 1 < len(src) else None

    return [
        sno,
        cell(SRC_SURVEY_ID),
        cell(SRC_OWNER),
        cell(SRC_FATHER),
        cell(SRC_MOBILE),
        cell(SRC_WARD),
        cell(SRC_PARCEL),
        cell(SRC_UNIT),
        cell(SRC_CITY),
        cell(SRC_PIN),
        cell(SRC_HOUSE),
        cell(SRC_COLONY),
        cell(SRC_TAX_ZONE),
        cell(SRC_PROP_TYPE),
        cell(SRC_PROP_USE),
        cell(SRC_ROAD),
    ]


def map_floors(src: list[Any]) -> tuple[str, dict[str, tuple[float, float]], float, int]:
    """Returns Floors code, {group: (res, non_res)}, open_land, commercial_hits."""
    pairs: dict[str, tuple[float, float]] = {g: (0.0, 0.0) for g in FLOOR_GROUPS}
    active: dict[str, float] = {g: 0.0 for g in FLOOR_GROUPS}
    open_land = 0.0
    commercial_hits = 0

    for name, area_col, uf_col, ut_col in SRC_FLOOR_BLOCKS:
        area = parse_area(src[area_col - 1] if area_col - 1 < len(src) else None)
        if area is None:
            continue
        factor = src[uf_col - 1] if uf_col - 1 < len(src) else None
        utype = src[ut_col - 1] if ut_col - 1 < len(src) else None

        if name == "Open Land":
            open_land = area
            continue

        residential = is_residential(factor, utype)
        if not residential:
            commercial_hits += 1
        if residential:
            pairs[name] = (area, 0.0)
        else:
            pairs[name] = (0.0, area)
        if area > 0:
            active[name] = area

    pairs["Seventh Floor"] = (0.0, 0.0)
    return compute_floors_code(active), pairs, open_land, commercial_hits


def write_row(ws, excel_row: int, src: list[Any], sno: int) -> int:
    identity = map_identity(src, sno)
    for c, value in enumerate(identity, start=1):
        ws.cell(excel_row, c, value)

    code, pairs, open_land, commercial_hits = map_floors(src)
    ws.cell(excel_row, OUT_FLOORS, code)

    col = OUT_MATRIX_START
    floor_sum = 0.0
    ground_sum = 0.0
    for group in FLOOR_GROUPS:
        res, non = pairs[group]
        ws.cell(excel_row, col, res)
        ws.cell(excel_row, col + 1, non)
        floor_sum += res + non
        if group == "Ground Floor":
            ground_sum = res + non
        col += 2

    ws.cell(excel_row, OUT_OPEN, open_land)
    ws.cell(excel_row, OUT_OPEN + 1, 0)

    plot = area_or_zero(src[SRC_PLOT - 1] if SRC_PLOT - 1 < len(src) else None)
    plinth = parse_area(src[SRC_PLINTH - 1] if SRC_PLINTH - 1 < len(src) else None)
    built = parse_area(src[SRC_BUILT - 1] if SRC_BUILT - 1 < len(src) else None)

    if plinth is None or plinth == 0:
        plinth = ground_sum if ground_sum > 0 else 0.0
    if built is None or built == 0:
        built = floor_sum

    ws.cell(excel_row, OUT_PLOT, plot)
    ws.cell(excel_row, OUT_PLINTH, plinth)
    ws.cell(excel_row, OUT_BUILT, built)
    return commercial_hits


def set_widths(ws) -> None:
    widths = {
        1: 6,
        2: 26,
        3: 18,
        4: 18,
        5: 12,
        6: 20,
        7: 10,
        8: 10,
        9: 18,
        10: 10,
        11: 10,
        12: 14,
        13: 12,
        14: 14,
        15: 12,
        16: 10,
        17: 10,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(OUT_MATRIX_START, OUT_LAST + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11


def verify(ws, commercial_total: int) -> None:
    n = ws.max_row - 4
    print(f"Data rows: {n}")
    print(f"Merges: {len(list(ws.merged_cells.ranges))}")
    print(f"Freeze: {ws.freeze_panes}")
    print(f"Commercial->Non-Res: {commercial_total}")

    prev = (-1, -1.0, "")
    violations = 0
    na_tail = 0
    first_floor_positive = 0
    gf_like = 0
    open_p = 0

    for r in range(5, ws.max_row + 1):
        sno = ws.cell(r, 1).value
        if sno != r - 4:
            print(f"SN mismatch row {r}: {sno}")
            break
        key = parcel_sort_key(ws.cell(r, 7).value)
        if key < prev:
            violations += 1
        if key[0] == 1:
            na_tail += 1
        prev = key

        f1 = float(ws.cell(r, 22).value or 0) + float(ws.cell(r, 23).value or 0)
        if f1 > 0:
            first_floor_positive += 1
            code = str(ws.cell(r, OUT_FLOORS).value or "")
            if "F1" in code or "F2" in code or "F3" in code:
                gf_like += 1

        code = str(ws.cell(r, OUT_FLOORS).value or "")
        open_v = float(ws.cell(r, OUT_OPEN).value or 0)
        built_sum = sum(float(ws.cell(r, c).value or 0) for c in range(18, 36))
        if code == "P" and open_v > 0 and built_sum == 0:
            open_p += 1

    print(f"Parcel sort violations: {violations}")
    print(f"N/A parcels at end: {na_tail}")
    print(f"Rows with First Floor area > 0: {first_floor_positive}")
    print(f"Those with F1/F2/F3 in Floors code: {gf_like}")
    print(f"Open-land-only P rows: {open_p}")

    print("\nSpot-check (First Floor > 0):")
    shown = 0
    for r in range(5, ws.max_row + 1):
        f1 = float(ws.cell(r, 22).value or 0) + float(ws.cell(r, 23).value or 0)
        if f1 <= 0:
            continue
        print(
            f"  SN={ws.cell(r,1).value} parcel={ws.cell(r,7).value} "
            f"Floors={ws.cell(r,17).value} G={ws.cell(r,20).value}/{ws.cell(r,21).value} "
            f"F1={ws.cell(r,22).value}/{ws.cell(r,23).value} "
            f"F2={ws.cell(r,24).value}/{ws.cell(r,25).value}"
        )
        shown += 1
        if shown >= 10:
            break

    print("\nNon-Res samples:")
    shown = 0
    for r in range(5, ws.max_row + 1):
        for c in range(19, 36, 2):
            v = float(ws.cell(r, c).value or 0)
            if v > 0:
                print(
                    f"  SN={ws.cell(r,1).value} parcel={ws.cell(r,7).value} "
                    f"Floors={ws.cell(r,17).value} col{c}={v}"
                )
                shown += 1
                break
        if shown >= 5:
            break


def main() -> int:
    repo = Path(__file__).resolve().parents[2]
    etm = repo / "ETM"
    default_zip = Path(r"C:\Users\sikar\Downloads\survey-data-district-ETA-wards.zip")
    default_src = etm / "1-Ward-1-Etah.xlsx"

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        type=Path,
        default=default_src if default_src.exists() else default_zip,
    )
    parser.add_argument("--zip-entry", default="ETM/1-Ward-1-Etah.xlsx")
    parser.add_argument(
        "--reference",
        type=Path,
        default=Path(r"C:\Users\sikar\Downloads\WARD 1.xlsx"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=etm / "1-Ward-1-Etah-floor-redesign.xlsx",
    )
    args = parser.parse_args()

    src_path = resolve_input(args.input, args.zip_entry)
    etm.mkdir(parents=True, exist_ok=True)
    local_src = etm / "1-Ward-1-Etah.xlsx"
    if src_path.resolve() != local_src.resolve():
        local_src.write_bytes(src_path.read_bytes())
        src_path = local_src

    print(f"Loading: {src_path}")
    src_wb = load_workbook(src_path, data_only=False)
    src_ws = src_wb.active
    rows = load_source_rows(src_ws)
    rows_sorted = sorted(rows, key=lambda r: (parcel_sort_key(r[SRC_PARCEL - 1]),))
    print(f"Source rows: {len(rows_sorted)}")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Survey Data"
    build_header(out_ws)

    commercial_total = 0
    for i, src in enumerate(rows_sorted):
        commercial_total += write_row(out_ws, 5 + i, src, i + 1)

    out_ws.freeze_panes = "A5"
    out_ws.auto_filter.ref = f"A4:{get_column_letter(OUT_LAST)}{4 + len(rows_sorted)}"
    set_widths(out_ws)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(args.output)
    print(f"Wrote: {args.output}")
    verify(out_ws, commercial_total)
    if args.reference.exists():
        print(f"Reference: {args.reference}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
